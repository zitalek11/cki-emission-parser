from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from tests.helpers import build_job, field_spec, mini_set

from cki_emission_parser.cli import main
from cki_emission_parser.extraction.llm import NullLlmProvider
from cki_emission_parser.extraction.pipeline import extract_job
from cki_emission_parser.models.types import Evidence, ExtractReport, FieldResult, UnmappedFact
from cki_emission_parser.output import report_to_dict, write_extract_output
from cki_emission_parser.output.excel import write_extract_excel
from cki_emission_parser.output.html_review import write_review_html


def _sample_set():
    return mini_set(
        field_spec(id="issuer.inn", nrd_path="issuer.inn", title="ИНН эмитента", type="string"),
        field_spec(id="cfi", nrd_path="cfi", title="Код CFI", type="string"),
        field_spec(
            id="state_reg_date",
            nrd_path="state_reg_date",
            title="Дата регистрации выпуска",
            type="date",
        ),
        field_spec(
            id="bond.quantity",
            nrd_path="bond.quantity",
            title="Количество облигаций",
            type="integer",
        ),
        field_spec(),
    )


def _sample_report() -> ExtractReport:
    return ExtractReport(
        job_id="job-output",
        instrument_class="bond_exchange",
        llm_used=False,
        fields=[
            FieldResult(
                field="issuer.inn",
                raw_value="7707083893",
                normalized_value="7707083893",
                status="confirmed",
                evidence=[Evidence(source_id="src_inn", quote="ИНН 7707083893", page=1)],
                canonical_source="src_inn",
                review_decision="accepted",
            ),
            FieldResult(
                field="cfi",
                raw_value="DBFUFB",
                normalized_value="DBFUFB",
                status="confirmed",
                evidence=[Evidence(source_id="src_cfi", quote="CFI: DBFUFB", page=2)],
                canonical_source="src_cfi",
                review_decision="accepted",
            ),
            FieldResult(
                field="state_reg_date",
                raw_value={
                    "date_kind": "rule",
                    "raw_text": "3-й рабочий день с даты начала размещения",
                    "normalized": None,
                },
                normalized_value={
                    "date_kind": "rule",
                    "raw_text": "3-й рабочий день с даты начала размещения",
                    "normalized": None,
                },
                status="confirmed",
                evidence=[
                    Evidence(
                        source_id="src_date",
                        quote="Дата: 3-й рабочий день с даты начала размещения",
                        page=3,
                    )
                ],
                canonical_source="src_date",
                review_decision="accepted",
            ),
            FieldResult(
                field="bond.quantity",
                raw_value=12,
                normalized_value=12,
                status="confirmed",
                evidence=[Evidence(source_id="src_qty", quote="Количество: 12 штук", page=4)],
                canonical_source="src_qty",
                review_decision="accepted",
            ),
            FieldResult(
                field="issuer.name_full",
                raw_value=["ПАО МТС", "ПАО «Мобильные ТелеСистемы»"],
                status="conflict",
                evidence=[
                    Evidence(
                        source_id="src_a",
                        quote="Полное фирменное наименование: ПАО МТС",
                        page=1,
                    ),
                    Evidence(
                        source_id="src_b",
                        quote="Полное фирменное наименование: ПАО «Мобильные ТелеСистемы»",
                        page=5,
                    ),
                ],
                review_decision="review_required",
                reason="документы расходятся",
            ),
        ],
        unmapped_facts=[
            UnmappedFact(
                label="Барьер",
                value="70%",
                source=Evidence(source_id="src_bar", quote="Барьер: 70%", page=8),
            )
        ],
    )


def _row_for(sheet, field_id: str):
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == field_id:
            return row
    raise AssertionError(field_id)


def test_extract_report_json_has_review_queue_not_fragments() -> None:
    report = _sample_report()
    payload = report_to_dict(report)
    assert "fields" in payload
    assert "status_counts" in payload
    assert payload["review_queue"]["conflict_count"] == 1
    assert "issuer.name_full" in payload["review_queue"]["field_ids"]
    assert "fragments" not in payload
    assert "retrieval" not in payload


def test_excel_keeps_inn_cfi_and_date_rules_as_text(tmp_path: Path) -> None:
    path = tmp_path / "report.xlsx"
    write_extract_excel(_sample_report(), path, extract_set=_sample_set())
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Поля", "На ревью", "Несопоставленные"]
    fields = workbook["Поля"]

    inn = _row_for(fields, "issuer.inn")[5]
    assert inn.value == "7707083893"
    assert isinstance(inn.value, str)
    assert inn.number_format == "@"
    assert not isinstance(inn.value, (int, float, datetime))

    cfi = _row_for(fields, "cfi")[5]
    assert cfi.value == "DBFUFB"
    assert isinstance(cfi.value, str)
    assert cfi.number_format == "@"

    date_cell = _row_for(fields, "state_reg_date")[5]
    assert "правило:" in date_cell.value
    assert "3-й рабочий день" in date_cell.value
    assert isinstance(date_cell.value, str)
    assert date_cell.number_format == "@"
    assert not isinstance(date_cell.value, datetime)

    quantity_text = _row_for(fields, "bond.quantity")[5]
    quantity_number = _row_for(fields, "bond.quantity")[7]
    assert quantity_text.value == "12"
    assert isinstance(quantity_text.value, str)
    assert quantity_text.number_format == "@"
    assert not isinstance(quantity_text.value, datetime)
    assert quantity_number.value == 12
    assert isinstance(quantity_number.value, int)

    review = workbook["На ревью"]
    assert _row_for(review, "issuer.name_full")[2].value == "Конфликт"
    unmapped = workbook["Несопоставленные"]
    assert unmapped.cell(2, 1).value == "Барьер"
    assert unmapped.cell(2, 2).value == "70%"
    assert unmapped.cell(2, 2).number_format == "@"


def test_html_review_shows_conflict_quotes_in_russian(tmp_path: Path) -> None:
    path = tmp_path / "review.html"
    write_review_html(_sample_report(), path, extract_set=_sample_set())
    text = path.read_text(encoding="utf-8")
    assert "<html lang=\"ru\">" in text
    assert "Конфликт" in text
    assert "Полное фирменное наименование: ПАО МТС" in text
    assert "ПАО «Мобильные ТелеСистемы»" in text
    assert "Несопоставленные факты" in text
    assert "Барьер" in text
    assert "70%" in text


def test_write_extract_output_picks_format_from_suffix(tmp_path: Path) -> None:
    report = _sample_report()
    json_path = tmp_path / "out.json"
    html_path = tmp_path / "out.html"
    xlsx_path = tmp_path / "out.xlsx"
    assert write_extract_output(report, json_path, extract_set=_sample_set()) == "json"
    assert write_extract_output(report, html_path, extract_set=_sample_set()) == "html"
    assert write_extract_output(report, xlsx_path, extract_set=_sample_set()) == "xlsx"
    assert "review_queue" in json_path.read_text(encoding="utf-8")
    assert "retrieval" not in json_path.read_text(encoding="utf-8")


def test_extract_cli_writes_xlsx(digital_pdf: Path, tmp_path: Path, monkeypatch) -> None:
    monkeypatch.delenv("CKI_LLM_API_KEY", raising=False)
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    out = tmp_path / "cli-report.xlsx"
    code = main([str(digital_pdf), "--extract", "--instrument", "bond_exchange", "--out", str(out)])
    assert code == 0
    workbook = load_workbook(out)
    assert "Поля" in workbook.sheetnames
    assert "На ревью" in workbook.sheetnames
    assert "Несопоставленные" in workbook.sheetnames


def test_pipeline_json_still_omits_fragments() -> None:
    job = build_job("Полное фирменное наименование эмитента: ПАО «Пример»")
    report = extract_job(
        job,
        provider=NullLlmProvider(),
        extract_set=mini_set(field_spec()),
        instrument_class="bond_exchange",
    )
    payload = report_to_dict(report)
    assert "fragments" not in payload
    assert "retrieval" not in payload
    assert "review_queue" in payload
