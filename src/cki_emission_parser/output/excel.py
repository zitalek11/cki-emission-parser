from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from cki_emission_parser.models.schema import ExtractSet, FieldSpec
from cki_emission_parser.models.types import ExtractReport, FieldResult
from cki_emission_parser.output.review import (
    format_display_value,
    needs_review,
    primary_page,
    quotes_joined,
    review_label_ru,
    status_label_ru,
)
from cki_emission_parser.schema import load_extract_set

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
_CONFLICT_FILL = PatternFill("solid", fgColor="F8D7DA")
_WRAP = Alignment(wrap_text=True, vertical="top")

_HEADERS = [
    "Поле",
    "Название",
    "Статус",
    "Код статуса",
    "Решение ревью",
    "Значение (текст)",
    "Нормализованное (текст)",
    "Число",
    "Цитаты",
    "source_id",
    "Страница",
    "Причина",
]


def write_extract_excel(
    report: ExtractReport,
    path: Path,
    *,
    extract_set: ExtractSet | None = None,
) -> None:
    extract_set = extract_set or load_extract_set()
    specs = {spec.id: spec for spec in extract_set.fields}
    workbook = Workbook()
    fields_sheet = workbook.active
    fields_sheet.title = "Поля"
    _write_fields_sheet(fields_sheet, report.fields, specs, review_only=False)

    review_sheet = workbook.create_sheet("На ревью")
    review_fields = [field for field in report.fields if needs_review(field)]
    _write_fields_sheet(review_sheet, review_fields, specs, review_only=True)

    unmapped_sheet = workbook.create_sheet("Несопоставленные")
    _write_unmapped_sheet(unmapped_sheet, report)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_fields_sheet(
    sheet: Worksheet,
    fields: list[FieldResult],
    specs: dict[str, FieldSpec],
    *,
    review_only: bool,
) -> None:
    _write_header(sheet, _HEADERS)
    for row_idx, field in enumerate(fields, start=2):
        spec = specs.get(field.field)
        title = spec.title if spec else field.field
        values = [
            field.field,
            title,
            status_label_ru(field.status),
            field.status,
            review_label_ru(field.review_decision),
            format_display_value(field.raw_value),
            format_display_value(field.normalized_value),
            _numeric_value(spec, field),
            quotes_joined(field),
            field.canonical_source or "",
            primary_page(field),
            field.reason or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if col_idx == 8:
                _set_number_or_empty(cell, value)
            elif col_idx == 11:
                _set_page(cell, value)
            else:
                _set_text(cell, value)
            cell.alignment = _WRAP
            if field.status == "conflict":
                cell.fill = _CONFLICT_FILL
            elif review_only or needs_review(field):
                cell.fill = _REVIEW_FILL
    _autosize(sheet, _HEADERS)
    _set_filter(sheet, _HEADERS)


def _write_unmapped_sheet(sheet: Worksheet, report: ExtractReport) -> None:
    headers = ["Подпись", "Значение", "Цитата", "source_id", "Страница", "Возможное поле"]
    _write_header(sheet, headers)
    for row_idx, fact in enumerate(report.unmapped_facts, start=2):
        values = [
            fact.label,
            fact.value,
            fact.source.quote,
            fact.source.source_id,
            fact.source.page,
            fact.possible_field or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if col_idx == 5:
                _set_page(cell, value)
            else:
                _set_text(cell, value)
            cell.alignment = _WRAP
    _autosize(sheet, headers)
    _set_filter(sheet, headers)


def _write_header(sheet: Worksheet, headers: list[str]) -> None:
    for col_idx, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"


def _set_text(cell, value: object) -> None:
    cell.number_format = "@"
    if isinstance(value, datetime):
        cell.value = value.isoformat(sep=" ", timespec="seconds")
        return
    cell.value = "" if value is None else str(value)


def _set_number_or_empty(cell, value: object) -> None:
    if value is None or isinstance(value, datetime):
        cell.value = None
        cell.number_format = "General"
        return
    cell.value = value
    cell.number_format = "0"


def _set_page(cell, value: object) -> None:
    if isinstance(value, int) and not isinstance(value, bool):
        cell.value = value
        cell.number_format = "0"
        return
    _set_text(cell, value if value is not None else "")


def _numeric_value(spec: FieldSpec | None, field: FieldResult) -> int | float | None:
    if spec is None or spec.type not in {"integer", "money"}:
        return None
    value = field.normalized_value if field.normalized_value is not None else field.raw_value
    if isinstance(value, dict) and "amount" in value:
        amount = value.get("amount")
        return amount if isinstance(amount, (int, float)) and not isinstance(amount, bool) else None
    if isinstance(value, bool) or isinstance(value, datetime):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def _autosize(sheet: Worksheet, headers: list[str]) -> None:
    widths = {
        1: 28,
        2: 36,
        3: 18,
        4: 14,
        5: 16,
        6: 32,
        7: 32,
        8: 12,
        9: 48,
        10: 18,
        11: 12,
        12: 36,
    }
    for index, _header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(index, 24)
    sheet.row_dimensions[1].height = 22


def _set_filter(sheet: Worksheet, headers: list[str]) -> None:
    last = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"
