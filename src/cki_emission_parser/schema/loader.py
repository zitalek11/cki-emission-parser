from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import load_workbook

from cki_emission_parser.models.schema import CatalogField, ExtractSet, FieldSpec, NrdCatalog
from cki_emission_parser.paths import config_dir


def load_extract_set(path: Path | None = None) -> ExtractSet:
    path = path or config_dir() / "schema" / "extract_set.yaml"
    raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    fields = [FieldSpec.model_validate(item) for item in raw.get("fields", [])]
    return ExtractSet(
        version=int(raw.get("version", 1)),
        instrument_classes=list(raw.get("instrument_classes", [])),
        fields=fields,
    )


def load_nrd_catalog(path: Path | None = None) -> NrdCatalog:
    path = path or config_dir() / "schema" / "nrd_catalog.xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    fields: list[CatalogField] = []
    block: str | None = None
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        name, description = row
        if not name or not isinstance(name, str):
            continue
        if name.startswith("Информационный"):
            block = name
            continue
        if name in {"Параметр", "Ссылки", "API NSD"}:
            continue
        fields.append(
            CatalogField(
                name=name.strip(),
                description=description.strip() if isinstance(description, str) else None,
                block=block,
            )
        )
    wb.close()
    return NrdCatalog(fields=fields)


def load_document_types(path: Path | None = None) -> dict:
    path = path or config_dir() / "schema" / "document_types.yaml"
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def load_normalization(path: Path | None = None) -> dict:
    path = path or config_dir() / "schema" / "normalization.yaml"
    return yaml.safe_load(path.read_text(encoding="utf-8"))
